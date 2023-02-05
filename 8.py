# import pandas as pd
#
# # 使用 ExcelFile ，通过将 xls 或者 xlsx 路径传入，生成一个实例
# xlsx = pd.ExcelFile(r'C:\Users\14250\Desktop\ex.xlsx')
# print(type(xlsx))
# print(xlsx)


# import openpyxl
# wb = openpyxl.load_workbook(r'C:\Users\14250\Desktop\ex.xlsx')
# sheet = wb['总表']
# print(cell(K, 4).value)

# coding:utf-8
import openpyxl

workbook = openpyxl.load_workbook(r'C:\Users\14250\Desktop\ex.xlsx')  # 打开工作簿
# print('1.', workbook.sheetnames)  # 获取每个工作表名字
# sheet=workbook.active      #打开活动表  只有一个sheet的时候去使用，获取指定的工作表
sheet = workbook['总表']  # 获取指定的工作表
# 获取某个的单元格
# cell = sheet['A4']
# print('3.', cell.value)  # 打印格式为单个有value值
# 获取一系列的单元格
cells = sheet['F48:F53']  # 起始单元格
i = 48  # 起始行数
score = 0
items1 = ["K", "L", "O", "P", "Q", "V", "W", "X"]
items2 = ["R", "S", "T", "U"]
for cell in cells:
    score = 0
    print(sheet['F' + str(i)].value)
    for item in items1:
        if int(sheet[item + str(i)].value) > 9:
            score = score + 10
        elif int(sheet[item + str(i)].value) > 6:
            score = score + 9
        elif int(sheet[item + str(i)].value) > 3:
            score = score + 8
        elif int(sheet[item + str(i)].value) > 0:
            score = score + 7
        elif int(sheet[item + str(i)].value) == 0:
            score = score + 6
        else:
            print("error")
    for item in items2:
        if int(sheet[item + str(i)].value) > 4:
            score = score + 5
        elif int(sheet[item + str(i)].value) > 0:
            score = score + 4
        elif int(sheet[item + str(i)].value) == 0:
            score = score + 3
        else:
            print("error")
    print(score)
    i = i + 1
